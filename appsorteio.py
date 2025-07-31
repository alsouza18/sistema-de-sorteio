import sys
import os
import random
import json
from datetime import datetime
from pathlib import Path
from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QWidget, QTabWidget,
    QPushButton, QLabel, QTextEdit, QFileDialog, QInputDialog,
    QHBoxLayout, QComboBox, QGroupBox, QMessageBox
)
from PyQt6.QtCore import Qt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg
from matplotlib.figure import Figure
from openpyxl import load_workbook, Workbook

class SorteadorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Sorteador Avançado v4.4 - Com Colocação")
        self.setGeometry(100, 100, 900, 700)
        
        self.historico = []
        self.dados_planilha = []
        self.colunas_disponiveis = []
        self.coluna_atual = 'A'
        self.arquivo_excel = None
        self.resultados = None
        
        self.configurar_ui()
        self.carregar_historico()

    def configurar_ui(self):
        tabs = QTabWidget()
        tabs.addTab(self.criar_aba_sorteio(), "🎲 Sorteio")
        tabs.addTab(self.criar_aba_historico(), "📜 Histórico")
        tabs.addTab(self.criar_aba_graficos(), "📊 Gráficos")
        self.setCentralWidget(tabs)

    def criar_aba_sorteio(self):
        tab = QWidget()
        layout = QVBoxLayout()
        
        grupo_excel = QGroupBox("Configuração do Excel")
        layout_excel = QVBoxLayout()
        
        self.btn_carregar = QPushButton("📂 Carregar Excel")
        self.btn_carregar.clicked.connect(self.carregar_excel)
        layout_excel.addWidget(self.btn_carregar)
        
        self.label_coluna = QLabel("Coluna para sorteio:")
        self.combo_coluna = QComboBox()
        self.combo_coluna.currentTextChanged.connect(self.mudar_coluna)
        layout_excel.addWidget(self.label_coluna)
        layout_excel.addWidget(self.combo_coluna)
        
        self.label_classificacao = QLabel("Classificação (opcional):")
        self.combo_classificacao = QComboBox()
        layout_excel.addWidget(self.label_classificacao)
        layout_excel.addWidget(self.combo_classificacao)
        
        grupo_excel.setLayout(layout_excel)
        layout.addWidget(grupo_excel)
        
        grupo_operacoes = QGroupBox("Operações")
        layout_operacoes = QVBoxLayout()
        
        botoes_layout = QHBoxLayout()
        
        self.btn_sortear = QPushButton("🔀 Sortear Itens")
        self.btn_sortear.clicked.connect(self.executar_sorteio)
        botoes_layout.addWidget(self.btn_sortear)
        
        self.btn_colocacao = QPushButton("🥇 Sortear com Colocação")
        self.btn_colocacao.clicked.connect(self.sortear_com_colocacao)
        botoes_layout.addWidget(self.btn_colocacao)
        
        self.btn_grupos = QPushButton("👥 Formar Grupos")
        self.btn_grupos.clicked.connect(self.criar_grupos)
        botoes_layout.addWidget(self.btn_grupos)
        
        layout_operacoes.addLayout(botoes_layout)
        
        botoes_layout2 = QHBoxLayout()
        
        self.btn_classificar = QPushButton("🏷️ Sortear por Classificação")
        self.btn_classificar.clicked.connect(self.sortear_por_classificacao)
        botoes_layout2.addWidget(self.btn_classificar)
        
        self.btn_exportar = QPushButton("💾 Exportar Resultados")
        self.btn_exportar.clicked.connect(self.exportar_resultados)
        botoes_layout2.addWidget(self.btn_exportar)
        
        layout_operacoes.addLayout(botoes_layout2)
        grupo_operacoes.setLayout(layout_operacoes)
        layout.addWidget(grupo_operacoes)
        
        self.label_resultado = QTextEdit()
        self.label_resultado.setReadOnly(True)
        layout.addWidget(self.label_resultado)
        
        tab.setLayout(layout)
        return tab

    def criar_aba_historico(self):
        tab = QWidget()
        layout = QVBoxLayout()
        self.label_historico = QTextEdit()
        self.label_historico.setReadOnly(True)
        layout.addWidget(self.label_historico)
        
        self.btn_limpar = QPushButton("🧹 Limpar Histórico")
        self.btn_limpar.clicked.connect(self.limpar_historico)
        layout.addWidget(self.btn_limpar)
        
        tab.setLayout(layout)
        return tab

    def criar_aba_graficos(self):
        tab = QWidget()
        layout = QVBoxLayout()
        self.figure = Figure(figsize=(8, 5))
        self.canvas = FigureCanvasQTAgg(self.figure)
        layout.addWidget(self.canvas)
        tab.setLayout(layout)
        return tab

    def carregar_excel(self):
        try:
            diretorio_inicial = str(Path.home() / "Documents")
            
            arquivo, _ = QFileDialog.getOpenFileName(
                self, 
                "Abrir Excel", 
                diretorio_inicial,
                "Planilhas (*.xlsx *.xls);;Todos os arquivos (*)"
            )
            
            if not arquivo:
                return
            
            arquivo = os.path.normpath(arquivo)
            
            if not os.path.exists(arquivo):
                raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")
                
            if not os.access(arquivo, os.R_OK):
                raise PermissionError(f"Sem permissão para ler o arquivo: {arquivo}")
            
            self.btn_carregar.setText("Carregando...")
            QApplication.processEvents()
            
            with open(arquivo, 'rb') as f:
                if f.read(4) != b'\x50\x4B\x03\x04':
                    raise ValueError("O arquivo selecionado não é um Excel válido")
            
            wb = load_workbook(arquivo)
            planilha = wb.active
            self.colunas_disponiveis = []
            
            for col in planilha.iter_cols():
                if any(cell.value for cell in col[1:]):
                    self.colunas_disponiveis.append(col[0].column_letter)
            
            if not self.colunas_disponiveis:
                raise ValueError("Nenhuma coluna com dados encontrada")
            
            self.combo_coluna.clear()
            self.combo_coluna.addItems(self.colunas_disponiveis)
            self.arquivo_excel = arquivo
            self.processar_dados_excel()
            self.gerar_grafico()
            
            self.label_resultado.setText(f"✅ Planilha carregada com sucesso!\n"
                                      f"Local: {arquivo}\n"
                                      f"{len(self.dados_planilha)} itens encontrados.")
            
        except Exception as e:
            QMessageBox.critical(self, "Erro", f"Não foi possível carregar o arquivo:\n{str(e)}")
            self.label_resultado.setText(f"❌ Erro ao carregar arquivo:\n{str(e)}")
        finally:
            self.btn_carregar.setText("📂 Carregar Excel")

    def processar_dados_excel(self):
        try:
            wb = load_workbook(self.arquivo_excel)
            planilha = wb.active
            self.dados_planilha = []
            classificacoes = set()
            
            for cell in planilha[self.coluna_atual][1:]:
                if cell.value:
                    item = {
                        'nome': str(cell.value).strip(),
                        'classificacao': str(planilha['B'][cell.row-1].value).strip() 
                            if 'B' in self.colunas_disponiveis and planilha['B'][cell.row-1].value 
                            else ''
                    }
                    self.dados_planilha.append(item)
                    if item['classificacao']:
                        classificacoes.add(item['classificacao'])
            
            self.combo_classificacao.clear()
            self.combo_classificacao.addItem("(Sem classificação)")
            self.combo_classificacao.addItems(sorted(classificacoes))
            
        except Exception as e:
            raise ValueError(f"Erro ao processar coluna {self.coluna_atual}:\n{str(e)}")

    def mudar_coluna(self, coluna):
        self.coluna_atual = coluna
        if self.arquivo_excel:
            try:
                self.processar_dados_excel()
                self.label_resultado.setText(f"✅ Coluna {coluna} carregada com sucesso!")
            except Exception as e:
                self.label_resultado.setText(f"❌ {str(e)}")

    def executar_sorteio(self):
        if not self.dados_planilha:
            QMessageBox.warning(self, "Aviso", "Carregue um Excel primeiro!")
            return
            
        quantidade, ok = QInputDialog.getInt(
            self, "Quantidade", "Quantos itens sortear?",
            min=1, max=len(self.dados_planilha), value=3
        )
            
        if ok:
            try:
                nomes = [item['nome'] for item in self.dados_planilha]
                sorteados = random.sample(nomes, quantidade)
                resultado = "🎉 ITENS SORTEADOS:\n\n• " + "\n• ".join(sorteados)
                self.label_resultado.setText(resultado)
                
                self.resultados = {
                    'tipo': 'Sorteio', 
                    'coluna': self.coluna_atual, 
                    'itens': sorteados,
                    'data': datetime.now().strftime("%d/%m/%Y %H:%M")
                }
                
                self.historico.append({
                    "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "tipo": "Sorteio",
                    "quantidade": quantidade,
                    "coluna": self.coluna_atual,
                    "itens": sorteados
                })
                self.atualizar_historico()
                self.gerar_grafico()
                
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Falha ao sortear itens:\n{str(e)}")

    def sortear_com_colocacao(self):
        if not self.dados_planilha:
            QMessageBox.warning(self, "Aviso", "Carregue um Excel primeiro!")
            return
            
        quantidade, ok = QInputDialog.getInt(
            self, "Quantidade", "Quantos itens sortear com colocação?",
            min=1, max=len(self.dados_planilha), value=3
        )
        
        if not ok:
            return
            
        premiar, ok_premio = QInputDialog.getItem(
            self, "Premiação", "Deseja adicionar prêmios aos colocados?",
            ["Não", "Sim - Prêmios padrão", "Sim - Definir prêmios"], 0, False
        )
        
        if ok:
            try:
                nomes = [item['nome'] for item in self.dados_planilha]
                sorteados = random.sample(nomes, quantidade)
                
                premios = []
                if premiar == "Sim - Prêmios padrão":
                    premios = ["Ouro", "Prata", "Bronze"] + [f"Menção {i}" for i in range(4, quantidade+1)]
                elif premiar == "Sim - Definir prêmios":
                    premios = []
                    for i in range(quantidade):
                        premio, ok = QInputDialog.getText(
                            self, f"Prêmio {i+1}º lugar", 
                            f"Digite o prêmio para o {i+1}º lugar:"
                        )
                        if ok and premio:
                            premios.append(premio)
                
                resultado = "🏆 RESULTADO COM COLOCAÇÃO:\n\n"
                for posicao, item in enumerate(sorteados, 1):
                    linha = f"{posicao}º lugar: {item}"
                    if premios and posicao <= len(premios):
                        linha += f" - Prêmio: {premios[posicao-1]}"
                    resultado += linha + "\n"
                
                self.label_resultado.setText(resultado)
                
                self.resultados = {
                    'tipo': 'Sorteio com Colocação', 
                    'coluna': self.coluna_atual, 
                    'itens': sorteados,
                    'premios': premios if premios else None,
                    'data': datetime.now().strftime("%d/%m/%Y %H:%M")
                }
                
                self.historico.append({
                    "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "tipo": "Sorteio com Colocação",
                    "quantidade": quantidade,
                    "coluna": self.coluna_atual,
                    "premiados": bool(premios),
                    "itens": sorteados
                })
                
                self.atualizar_historico()
                self.gerar_grafico()
                
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Falha ao sortear itens:\n{str(e)}")

    def criar_grupos(self):
        if not self.dados_planilha:
            QMessageBox.warning(self, "Aviso", "Carregue um Excel primeiro!")
            return
        
        num_grupos, ok = QInputDialog.getInt(
            self, "Formar Grupos", "Em quantos grupos dividir?",
            min=1, max=len(self.dados_planilha), value=2
        )
        
        if ok:
            try:
                nomes = [item['nome'] for item in self.dados_planilha]
                random.shuffle(nomes)
                grupos = [[] for _ in range(num_grupos)]
                
                for i, item in enumerate(nomes):
                    grupos[i % num_grupos].append(item)
                
                resultado = "🏆 GRUPOS CRIADOS:\n\n"
                for i, grupo in enumerate(grupos, 1):
                    resultado += f"Grupo {i} ({len(grupo)} itens):\n• " + "\n• ".join(grupo) + "\n\n"
                
                self.label_resultado.setText(resultado)
                
                self.resultados = {
                    'tipo': 'Grupos', 
                    'coluna': self.coluna_atual, 
                    'grupos': grupos,
                    'data': datetime.now().strftime("%d/%m/%Y %H:%M")
                }
                
                self.historico.append({
                    "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "tipo": "Grupos",
                    "num_grupos": num_grupos,
                    "coluna": self.coluna_atual,
                    "itens_por_grupo": [len(g) for g in grupos]
                })
                self.atualizar_historico()
                self.gerar_grafico()
                
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Falha ao criar grupos:\n{str(e)}")

    def sortear_por_classificacao(self):
        if not self.dados_planilha:
            QMessageBox.warning(self, "Aviso", "Carregue um Excel primeiro!")
            return
            
        classificacao = self.combo_classificacao.currentText()
        if classificacao == "(Sem classificação)":
            self.executar_sorteio()
            return
            
        itens_classificacao = [item['nome'] for item in self.dados_planilha 
                             if item.get('classificacao') == classificacao]
        
        if not itens_classificacao:
            QMessageBox.warning(self, "Aviso", f"Nenhum item com classificação '{classificacao}'")
            return
            
        quantidade, ok = QInputDialog.getInt(
            self, "Quantidade", f"Quantos itens sortear da classificação '{classificacao}'?",
            min=1, max=len(itens_classificacao), value=1
        )
            
        if ok:
            try:
                sorteados = random.sample(itens_classificacao, quantidade)
                resultado = f"🏷️ Itens da classificação '{classificacao}':\n\n• " + "\n• ".join(sorteados)
                self.label_resultado.setText(resultado)
                
                self.resultados = {
                    'tipo': 'Sorteio por Classificação', 
                    'classificacao': classificacao, 
                    'itens': sorteados,
                    'data': datetime.now().strftime("%d/%m/%Y %H:%M")
                }
                
                self.historico.append({
                    "data": datetime.now().strftime("%d/%m/%Y %H:%M"),
                    "tipo": "Sorteio por Classificação",
                    "classificacao": classificacao,
                    "quantidade": quantidade,
                    "itens": sorteados
                })
                self.atualizar_historico()
                
            except Exception as e:
                QMessageBox.critical(self, "Erro", f"Falha ao sortear por classificação:\n{str(e)}")

    def exportar_resultados(self):
        if not self.resultados:
            QMessageBox.warning(self, "Aviso", "Nenhum resultado para exportar!")
            return
        
        try:
            diretorio_padrao = str(Path.home() / "Documents")
            nome_padrao = f"Resultados_Sorteio_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            caminho_padrao = os.path.join(diretorio_padrao, nome_padrao)
            
            arquivo, _ = QFileDialog.getSaveFileName(
                self, 
                "Salvar Resultados", 
                caminho_padrao,
                "Excel (*.xlsx);;Todos os arquivos (*)"
            )
            
            if not arquivo:
                return
            
            if not arquivo.lower().endswith('.xlsx'):
                arquivo += '.xlsx'
            
            os.makedirs(os.path.dirname(arquivo), exist_ok=True)
            
            wb = Workbook()
            planilha = wb.active
            
            planilha.append(["RESULTADOS DO SORTEIO"])
            planilha.append(["Data:", self.resultados.get('data', datetime.now().strftime("%d/%m/%Y %H:%M"))])
            
            if 'coluna' in self.resultados:
                planilha.append(["Coluna:", self.resultados['coluna']])
            if 'classificacao' in self.resultados:
                planilha.append(["Classificação:", self.resultados['classificacao']])
            
            planilha.append([])
            
            if self.resultados['tipo'] == 'Sorteio':
                planilha.append(["Itens Sorteados"])
                for item in self.resultados['itens']:
                    planilha.append([item])
            
            elif self.resultados['tipo'] == 'Sorteio com Colocação':
                planilha.append(["Itens Sorteados com Colocação"])
                for posicao, item in enumerate(self.resultados['itens'], 1):
                    linha = [f"{posicao}º lugar: {item}"]
                    if self.resultados.get('premios') and posicao <= len(self.resultados['premios']):
                        linha.append(f"Prêmio: {self.resultados['premios'][posicao-1]}")
                    planilha.append(linha)
            
            elif self.resultados['tipo'] == 'Grupos':
                planilha.append(["Grupos Criados"])
                for i, grupo in enumerate(self.resultados['grupos'], 1):
                    planilha.append([f"Grupo {i}"])
                    for item in grupo:
                        planilha.append([item])
                    planilha.append([])
            
            elif self.resultados['tipo'] == 'Sorteio por Classificação':
                planilha.append([f"Itens da Classificação: {self.resultados['classificacao']}"])
                for item in self.resultados['itens']:
                    planilha.append([item])
            
            wb.save(arquivo)
            QMessageBox.information(
                self, 
                "Sucesso", 
                f"Resultados exportados com sucesso para:\n{arquivo}"
            )
            
        except PermissionError:
            QMessageBox.critical(
                self, 
                "Erro", 
                "Permissão negada. Feche o arquivo se estiver aberto ou escolha outro local."
            )
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Erro", 
                f"Falha ao exportar resultados:\n{str(e)}"
            )

    def carregar_historico(self):
        try:
            caminho_historico = Path.home() / "Documents" / "Sorteador_historico.json"
            
            if caminho_historico.exists():
                with open(caminho_historico, "r", encoding='utf-8') as f:
                    dados = json.load(f)
                    if isinstance(dados, list):
                        self.historico = dados
                    else:
                        self.historico = []
            
            self.atualizar_historico()
            
        except json.JSONDecodeError:
            QMessageBox.warning(
                self, 
                "Aviso", 
                "O arquivo de histórico está corrompido. Um novo será criado."
            )
            self.historico = []
        except Exception as e:
            QMessageBox.warning(
                self, 
                "Aviso", 
                f"Não foi possível carregar o histórico:\n{str(e)}"
            )
            self.historico = []

    def atualizar_historico(self):
        try:
            texto = "📅 Histórico de Sorteios:\n\n"
            
            if not self.historico:
                texto += "Nenhum registro no histórico."
            else:
                for i, item in enumerate(reversed(self.historico), 1):
                    texto += f"{i}. {item.get('data', 'Data desconhecida')} - "
                    
                    if item.get("tipo") == "Grupos":
                        texto += (
                            f"{item.get('num_grupos', '?')} grupos "
                            f"(Coluna: {item.get('coluna', '?')}, "
                            f"Itens por grupo: {item.get('itens_por_grupo', [])})\n"
                        )
                    elif item.get("tipo") == "Sorteio por Classificação":
                        texto += (
                            f"{item.get('quantidade', '?')} itens da classificação "
                            f"'{item.get('classificacao', 'Desconhecida')}'\n"
                        )
                    elif item.get("tipo") == "Sorteio com Colocação":
                        texto += (
                            f"{item.get('quantidade', '?')} itens sorteados com colocação "
                            f"(Coluna: {item.get('coluna', '?')})"
                        )
                        if item.get("premiados"):
                            texto += " [COM PRÊMIOS]"
                        texto += "\n"
                    else:
                        texto += (
                            f"{item.get('quantidade', '?')} itens sorteados "
                            f"(Coluna: {item.get('coluna', '?')})\n"
                        )
            
            self.label_historico.setText(texto)
            
        except Exception as e:
            self.label_historico.setText(f"Erro ao carregar histórico: {str(e)}")

    def limpar_historico(self):
        resposta = QMessageBox.question(
            self,
            "Confirmar",
            "Tem certeza que deseja limpar todo o histórico?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if resposta == QMessageBox.StandardButton.Yes:
            self.historico = []
            self.atualizar_historico()
            
            try:
                caminho_historico = Path.home() / "Documents" / "Sorteador_historico.json"
                if caminho_historico.exists():
                    caminho_historico.unlink()
            except Exception as e:
                QMessageBox.warning(
                    self, 
                    "Aviso", 
                    f"Não foi possível apagar o arquivo de histórico:\n{str(e)}"
                )

    def gerar_grafico(self):
        self.figure.clear()
        ax = self.figure.add_subplot(111)
        
        if not self.dados_planilha:
            ax.text(0.5, 0.5, "Sem dados para exibir", ha='center', va='center')
            ax.set_title("Nenhum dado carregado")
        else:
            classificacoes = [item.get('classificacao', '') 
                            for item in self.dados_planilha 
                            if item.get('classificacao', '')]
            
            if classificacoes:
                contagem_class = {}
                for cls in classificacoes:
                    contagem_class[cls] = contagem_class.get(cls, 0) + 1
                
                if len(contagem_class) > 1:
                    ax.pie(
                        contagem_class.values(),
                        labels=contagem_class.keys(),
                        autopct=lambda p: f'{p:.1f}%\n({int(p*sum(contagem_class.values()))/100})',
                        startangle=90,
                        textprops={'fontsize': 8}
                    )
                    ax.set_title(f"Distribuição por Classificação\nColuna {self.coluna_atual}")
                else:
                    self.plotar_distribuicao_letras(ax)
            else:
                self.plotar_distribuicao_letras(ax)
        
        self.canvas.draw()

    def plotar_distribuicao_letras(self, ax):
        letras = [nome[0].upper() for nome in 
                 [item['nome'] for item in self.dados_planilha if item['nome']]]
        contagem_letras = {}
        
        for letra in letras:
            contagem_letras[letra] = contagem_letras.get(letra, 0) + 1
        
        ax.bar(contagem_letras.keys(), contagem_letras.values())
        ax.set_xlabel("Letra Inicial")
        ax.set_ylabel("Quantidade")
        ax.set_title(f"Distribuição por Letra Inicial\nColuna {self.coluna_atual}")
        
        for i, v in enumerate(contagem_letras.values()):
            ax.text(i, v + 0.5, str(v), ha='center')

    def closeEvent(self, event):
        try:
            caminho_historico = Path.home() / "Documents" / "Sorteador_historico.json"
            
            with open(caminho_historico, "w", encoding='utf-8') as f:
                json.dump(self.historico, f, indent=4, ensure_ascii=False)
                
        except Exception as e:
            QMessageBox.warning(
                self, 
                "Aviso", 
                f"Não foi possível salvar o histórico:\n{str(e)}"
            )
        
        event.accept()

if __name__ == "__main__":
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        app.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        app.setAttribute(Qt.AA_UseHighDpiPixmaps, True)
    
    window = SorteadorApp()
    window.show()
    sys.exit(app.exec())